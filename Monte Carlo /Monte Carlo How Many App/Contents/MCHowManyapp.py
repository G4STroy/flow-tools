import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenPyXLImage
import os
# from PIL import Image, ImageTk, ImageResampling  # Commented out image-related imports
import tkinter as tk
from tkinter import filedialog, messagebox
import shutil
import sys
import traceback

class MonteCarloApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Monte Carlo How Many")

        # Define the base path for resources
        self.resources_dir = self.get_base_path()

        # Define the paths for resources
        self.initialize_resource_paths()

        # Load resources and setup UI after paths are set
        self.load_resources()
        self.setup_ui()

    def get_base_path(self):
        """Get the base path for resources."""
        if getattr(sys, 'frozen', False):
            # Running in a bundle
            return sys._MEIPASS
        else:
            return "/Users/troy.lightfoot/Github Projects/flow-tools/Monte Carlo /Monte Carlo How Many App/Contents/Resources"

    def initialize_resource_paths(self):
        """Initialize resource paths."""
        # self.image_path = os.path.join(self.resources_dir, 'monteimage.jpeg')  # Commented out image path
        self.template_file_path = os.path.join(self.resources_dir, 'Monte Carlo How Many.xlsx')
        self.instructions_file_path = os.path.join(self.resources_dir, 'Monte Carlo How Many User Guide.docx')

    def load_buttons(self, button_frame):
        """Load buttons within the button frame."""
        actions = [
            ("Download Template", self.download_template),
            ("Upload Modified Template", self.upload_template),
            ("Instructions", self.download_instructions)
        ]
        for text, command in actions:
            button = tk.Button(button_frame, text=text, command=command)
            button.pack(side=tk.LEFT, padx=5, pady=5)

    # def display_image(self, image_frame):
    #     # Attempt to load and display the image within the image frame
    #     try:
    #         image = Image.open(self.image_path)
    #         image = image.resize((100, 100), ImageResampling.LANCZOS)  # Updated this line
    #         img = ImageTk.PhotoImage(image)
    #         image_label = tk.Label(image_frame, image=img)
    #         image_label.image = img  # Keep a reference
    #         image_label.pack(side='top', pady=10, anchor='n')  # Adjusted this line
    #     except Exception as e:
    #         messagebox.showerror("Error", f"Failed to open image: {e}")
    #         traceback.print_exc()

    def setup_ui(self):
        # Create the UI elements here
        # Create a frame for the image
        # image_frame = tk.Frame(self.root)
        # image_frame.pack(fill=tk.BOTH, expand=True)

        # Create a frame for the buttons
        button_frame = tk.Frame(self.root)
        button_frame.pack(fill='x', expand=False)

        # self.display_image(image_frame)  # Commented out image display
        self.load_buttons(button_frame)

    def load_resources(self):
        # Define the paths for resources using `resources_dir`
        # self.image_path = os.path.join(self.resources_dir, 'monteimage.jpeg')
        self.template_file_path = os.path.join(self.resources_dir, 'Monte Carlo How Many.xlsx')
        self.instructions_file_path = os.path.join(self.resources_dir, 'Monte Carlo How Many User Guide.docx')

        # Debugging: check if the files actually exist at the specified paths
        for path in [self.template_file_path, self.instructions_file_path]:
            if os.path.isfile(path):
                print(f"Confirmed: The file exists at {path}")
            else:
                print(f"File not found at {path}")
        # Attempt to load and display the image
        # try:
        #     image = PILImage.open(self.image_path)
        #     img = ImageTk.PhotoImage(image)
        #     image_label = tk.Label(self.root, image=img)
        #     image_label.image = img  # Keep a reference so it's not garbage-collected
        #     image_label.pack(padx=10, pady=10)
        # except Exception as e:
        #     messagebox.showerror("Error", f"Failed to open image: {e}")
        #     traceback.print_exc()

        # Attempt to load the Excel file
        try:
            df = pd.read_excel(self.template_file_path)
            print("Excel file loaded successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Excel file: {e}")
            traceback.print_exc()

        # Attempt to open the User Guide
        try:
            with open(self.instructions_file_path, 'r') as f:
                print("User Guide loaded successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open User Guide: {e}")
            traceback.print_exc()    

    def check_file_existence(self):
        for file_path in [self.template_file_path, self.instructions_file_path]:
            print(f"{file_path} - {'found' if os.path.exists(file_path) else 'not found'}")

    def file_dialog_wrapper(self, dialog_func, dialog_title, file_types=None, is_save=False):
        self.root.withdraw()
        try:
            if dialog_func == filedialog.askdirectory:
                path = dialog_func(title=dialog_title)
            else:
                if file_types:
                    path = dialog_func(title=dialog_title, filetypes=file_types)
                else:
                    path = dialog_func(title=dialog_title)
        finally:
            self.root.deiconify()
            return path

    def download_template(self):
        destination_path = self.file_dialog_wrapper(filedialog.askdirectory, "Select Destination Directory")
        if destination_path:
            try:
                shutil.copy(self.template_file_path, os.path.join(destination_path, os.path.basename(self.template_file_path)))
                messagebox.showinfo("Success", "Template downloaded successfully.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to download the template: {e}")

    def upload_template(self):
        file_path = self.file_dialog_wrapper(filedialog.askopenfilename, "Select Template File",
                                             file_types=[("Excel files", "*.xlsx")])
        if file_path:
            self.process_template(file_path)

    def download_instructions(self):
        destination_path = self.file_dialog_wrapper(filedialog.askdirectory, "Select Destination Directory")
        if destination_path:
            try:
                shutil.copy(self.instructions_file_path, os.path.join(destination_path, os.path.basename(self.instructions_file_path)))
                messagebox.showinfo("Success", "Instructions downloaded successfully.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to download the instructions: {e}")

    def process_template(self, file_path):
        try:
            # Read the input data
            df_input = pd.read_excel(file_path, sheet_name='Data Input Sheet')

            # Validate the existence of necessary columns
            expected_columns = ['ID', 'Started', 'Done', 'Selected']
            for col in expected_columns:
                if col not in df_input.columns:
                    raise ValueError(f"Required column '{col}' is missing.")

            # Handle NaN values and incorrect types in critical columns
            df_input['ID'] = df_input['ID'].fillna('Unknown_ID')
            df_input['Started'] = pd.to_datetime(df_input['Started'], errors='coerce').dt.normalize()
            df_input['Done'] = pd.to_datetime(df_input['Done'], errors='coerce').dt.normalize()
            df_input['Selected'] = df_input['Selected'].fillna('no').str.lower()

            # Warn if any critical date columns have NaT after coercion
            if df_input['Started'].isna().any() or df_input['Done'].isna().any():
                messagebox.showwarning("Warning", "Some 'Started' or 'Done' dates are invalid and will be excluded from the simulation.")

            # Filter out rows with NaT in 'Started' or 'Done' column and where 'Selected' is not 'yes'
            df_selected = df_input.dropna(subset=['Started', 'Done'])
            df_selected = df_selected[df_selected['Selected'] == 'yes']

            # Proceed with processing...
            # Determine the earliest and latest dates from the selected data for the throughput calculation
            earliest_date = df_selected['Done'].min()
            latest_date = df_selected['Done'].max()

            # Create a date range for the throughput calculation
            throughput_dates = pd.date_range(start=earliest_date, end=latest_date, freq='D')

            # Calculate daily throughput within this range, including days with zero completions
            daily_throughput = df_selected.groupby('Done').size().reindex(throughput_dates, fill_value=0)

            # Check for NaN values in control parameters and handle them
            if daily_throughput.isna().any():
                daily_throughput.fillna(0, inplace=True)  # Replace NaNs with 0

            # Read simulation control parameters for the number of days to simulate
            df_control = pd.read_excel(file_path, sheet_name='Simulation Control Sheet')

            # Check for NaN values in control parameters and handle them
            if df_control['Total Simulations'].isna().any():
                messagebox.showerror("Error", "'Total Simulations' contains invalid data.")
                return
            if df_control['Days To Simulate'].isna().any():
                messagebox.showerror("Error", "'Days To Simulate' contains invalid data.")
                return

            num_simulations = 10000  # Hard coded number of simulations
            num_days_to_simulate = int(df_control.at[0, 'Days To Simulate'])

            # Perform simulations
            np.random.seed(42)  # For reproducibility
            simulation_results = []
            for _ in range(num_simulations):
                simulated_throughput = np.random.choice(daily_throughput, size=num_days_to_simulate, replace=True)
                simulation_results.append(np.sum(simulated_throughput))

            # Calculate the percentile values
            percentiles = [95, 85, 70, 50, 30]
            percentile_values = {f'{p}th Percentile': np.percentile(simulation_results, 100 - p) for p in percentiles}

            # Write Summary of Percentiles to Excel
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
                if 'Summary of Percentiles' in writer.book.sheetnames:
                    idx = writer.book.sheetnames.index('Summary of Percentiles')
                    writer.book.remove(writer.book.worksheets[idx])
                pd.DataFrame.from_dict(percentile_values, orient='index', columns=['Total Throughput']).to_excel(writer, sheet_name='Summary of Percentiles', index=True)

            # Generate a Histogram and Insert it into the Excel Workbook
            plt.figure(figsize=(10, 6))
            plt.hist(simulation_results, bins=50, alpha=0.75)
            plt.title('Histogram of Total Throughput per Simulation')
            histogram_path = os.path.join(self.resources_dir, 'simulation_totals_histogram.png')
            plt.savefig(histogram_path)
            plt.close()

            # Insert the histogram into the Excel workbook
            wb = load_workbook(file_path)
            if 'Histogram' in wb.sheetnames:
                wb.remove(wb['Histogram'])
            ws_histogram = wb.create_sheet('Histogram')
            img = OpenPyXLImage(histogram_path)
            ws_histogram.add_image(img, 'A1')
            wb.save(file_path)

            # Show message box to inform the user
            messagebox.showinfo("Monte Carlo Process Completed", "Monte Carlo simulation process completed. Results and histogram have been written to the Excel workbook.")

            print("Monte Carlo simulation process completed. Results and histogram have been written to the Excel workbook.")

        except ValueError as ve:
            messagebox.showerror("Error", f"Error processing template: {ve}")
            print(f"Error processing template: {ve}")
            pass
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")
            print(f"An error occurred: {e}")
            pass


def main():
    root = tk.Tk()
    app = MonteCarloApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
