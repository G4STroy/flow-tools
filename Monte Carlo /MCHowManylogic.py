import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os
from PIL import Image as PILImage

# Correct the file path
project_folder = 'flow-tools/Monte Carlo '
excel_file_name = 'Monte Carlo How Many.xlsx'
excel_file_path = os.path.join('/Users/troy.lightfoot/Github Projects', project_folder, excel_file_name)

# Read the input data
df_input = pd.read_excel(excel_file_path, sheet_name='Data Input Sheet')
df_selected = df_input[df_input['Selected'].str.lower() == 'yes'].copy()

# Convert 'Done' column to datetime and normalize to remove the time portion
df_selected['Done'] = pd.to_datetime(df_selected['Done']).dt.normalize()

# Determine the earliest and latest dates from the selected data for the throughput calculation
earliest_date = df_selected['Done'].min()
latest_date = df_selected['Done'].max()

# Create a date range for the throughput calculation
throughput_dates = pd.date_range(start=earliest_date, end=latest_date, freq='D')

# Calculate daily throughput within this range, including days with zero completions
daily_throughput = df_selected.groupby('Done').size().reindex(throughput_dates, fill_value=0)

# Read simulation control parameters for the number of days to simulate
df_control = pd.read_excel(excel_file_path, sheet_name='Simulation Control Sheet')
num_simulations = int(df_control.at[0, 'Total Simulations'])
num_days_to_simulate = int(df_control.at[0, 'Days To Simulate'])

# Perform simulations
np.random.seed(42)  # For reproducibility
simulation_results = []
for _ in range(num_simulations):
    simulated_throughput = np.random.choice(daily_throughput, size=num_days_to_simulate, replace=True)
    simulation_results.append(np.sum(simulated_throughput))

# Calculate the percentile values
percentiles = [95, 85, 70, 50, 30]
percentile_values = {f'{p}th Percentile': np.percentile(simulation_results, 100 - p) for p in percentiles}
df_percentiles = pd.DataFrame.from_dict(percentile_values, orient='index', columns=['Total Throughput'])

# Write Summary of Percentiles to Excel
with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a') as writer:
    if 'Summary of Percentiles' in writer.book.sheetnames:
        idx = writer.book.sheetnames.index('Summary of Percentiles')
        writer.book.remove(writer.book.worksheets[idx])
    df_percentiles.to_excel(writer, sheet_name='Summary of Percentiles', index=True)

# Generate a Histogram and Insert it into the Excel Workbook
plt.figure(figsize=(10, 6))
plt.hist(simulation_results, bins=50, alpha=0.75)
plt.title('Histogram of Total Throughput per Simulation')
plt.xlabel('Total Throughput')
plt.ylabel('Frequency')
plt.grid(True)
histogram_path = 'simulation_totals_histogram.png'
plt.savefig(histogram_path)
plt.close()

# Insert the histogram into the Excel workbook
wb = load_workbook(excel_file_path)
if 'Histogram' in wb.sheetnames:
    wb.remove(wb['Histogram'])
ws_histogram = wb.create_sheet('Histogram')
img = Image(histogram_path)
ws_histogram.add_image(img, 'A1')
wb.save(excel_file_path)

print("Monte Carlo simulation process completed. Results and histogram have been written to the Excel workbook.")
