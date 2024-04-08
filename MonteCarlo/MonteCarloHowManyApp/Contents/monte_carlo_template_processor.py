import pandas as pd
from tkinter import messagebox
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenPyXLImage
import os


class MonteCarloTemplateProcessor:
    def __init__(self, file_manager):
        self.file_manager = file_manager

    def process_template(self, file_path):
        try:
            df_input = pd.read_excel(file_path, sheet_name='Data Input Sheet')
            expected_columns = ['ID', 'Started', 'Done', 'Selected']
            for col in expected_columns:
                if col not in df_input.columns:
                    raise ValueError(f"Required column '{col}' is missing.")
            df_input['ID'] = df_input['ID'].fillna('Unknown_ID')
            df_input['Started'] = pd.to_datetime(df_input['Started'], errors='coerce').dt.normalize()
            df_input['Done'] = pd.to_datetime(df_input['Done'], errors='coerce').dt.normalize()
            df_input['Selected'] = df_input['Selected'].fillna('no').str.lower()
            if df_input['Started'].isna().any() or df_input['Done'].isna().any():
                raise ValueError("Some 'Started' or 'Done' dates are invalid and will be excluded from the simulation.")
            df_selected = df_input.dropna(subset=['Started', 'Done'])
            df_selected = df_selected[df_selected['Selected'] == 'yes']
            earliest_date = df_selected['Done'].min()
            latest_date = df_selected['Done'].max()
            throughput_dates = pd.date_range(start=earliest_date, end=latest_date, freq='D')
            daily_throughput = df_selected.groupby('Done').size().reindex(throughput_dates, fill_value=0)
            if daily_throughput.isna().any():
                daily_throughput.fillna(0, inplace=True)
            df_control = pd.read_excel(file_path, sheet_name='Simulation Control Sheet')
            if df_control['Total Simulations'].isna().any() or df_control['Days To Simulate'].isna().any():
                raise ValueError("'Total Simulations' or 'Days To Simulate' contains invalid data.")
            num_simulations = 10000
            num_days_to_simulate = int(df_control.at[0, 'Days To Simulate'])
            np.random.seed(42)
            simulation_results = []
            for _ in range(num_simulations):
                simulated_throughput = np.random.choice(daily_throughput, size=num_days_to_simulate, replace=True)
                simulation_results.append(np.sum(simulated_throughput))
            percentiles = [95, 85, 70, 50, 30]
            percentile_values = {f'{p}th Percentile': np.percentile(simulation_results, 100 - p) for p in percentiles}
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
                if 'Summary of Percentiles' in writer.book.sheetnames:
                    idx = writer.book.sheetnames.index('Summary of Percentiles')
                    writer.book.remove(writer.book.worksheets[idx])
                pd.DataFrame.from_dict(percentile_values, orient='index', columns=['Total Throughput']).to_excel(writer, sheet_name='Summary of Percentiles', index=True)
            plt.figure(figsize=(10, 6))
            plt.hist(simulation_results, bins=50, alpha=0.75)
            plt.title('Histogram of Total Throughput per Simulation')

            # Save the histogram in the same directory as the Excel file
            histogram_path = os.path.join(os.path.dirname(file_path), 'simulation_totals_histogram.png')

            plt.savefig(histogram_path)
            plt.close()
            wb = load_workbook(file_path)
            if 'Histogram' in wb.sheetnames:
                wb.remove(wb['Histogram'])
            ws_histogram = wb.create_sheet('Histogram')
            img = OpenPyXLImage(histogram_path)
            ws_histogram.add_image(img, 'A1')
            wb.save(file_path)
            messagebox.showinfo("Monte Carlo Process Completed", "Monte Carlo simulation process completed. Results and histogram have been written to the Excel workbook.")
            print("Monte Carlo simulation process completed. Results and histogram have been written to the Excel workbook.")
        except ValueError as ve:
            messagebox.showerror("Error", f"Error processing template: {ve}")
            print(f"Error processing template: {ve}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")
            print(f"An error occurred: {e}")